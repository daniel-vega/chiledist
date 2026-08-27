"""
tests/test_validation_election.py
====================================
Tests para chiledist.domain.data.tricel (import_proclamations/import_votes)
y chiledist.validation (ValidationReport/validate_election).

No requiere Excel reales de TRICEL/SERVEL ni shapefiles APC2023: los
tests construyen archivos .xlsx sintéticos en tmp_path reproduciendo la
forma exacta de DISTRITO-XX.xlsx determinada por inspección directa
(hojas ELECTOS/CANDIDATOS con filas de título antes del encabezado real
— ver chiledist/domain/data/tricel/__init__.py) y usan
commune_cut_map/pacto_map sintéticos como en tests/test_servel_candidates.py.
"""

from __future__ import annotations

import dataclasses
import os

import openpyxl
import pandas as pd
import pytest

from chiledist.domain.data import get_servel_dir, get_tricel_dir
from chiledist.domain.data.tricel import (
    ProclamationRecord,
    import_proclamations,
)
from chiledist.validation import ValidationReport, validate_election


# ──────────────────────────────────────────────────────────────────────────────
# Fixtures sintéticas — SERVEL (mismo patrón que test_servel_candidates.py)
# ──────────────────────────────────────────────────────────────────────────────

_FAKE_COMMUNE_CUT_MAP = {"SANTIAGO": 13101}


def _raw_servel_rows(distrito="Distrito 1"):
    """
    2 candidatos electos por SERVEL, cod_candidato terminado en 01/02 —
    corresponden a num_tricel=1/2 en la fixture TRICEL de abajo.
    """
    return [
        {
            "region": 13, "distrito": distrito, "comuna": "Santiago",
            "pacto": "Unidad por Chile", "subpacto": "A", "partido": "PS",
            "cod_candidato": 1001, "nombre_candidato": "Candidato Uno",
            "votos_preliminares": 1000, "electo_nominado": 1,
        },
        {
            "region": 13, "distrito": distrito, "comuna": "Santiago",
            "pacto": "Chile Grande y Unido", "subpacto": "B", "partido": "RN",
            "cod_candidato": 1002, "nombre_candidato": "Candidato Dos",
            "votos_preliminares": 500, "electo_nominado": 0,
        },
    ]


def _write_servel_excel(dir_path, rows, filename="PRELIMINARES_DIPUTADOS_DISTRITO_1.xlsx"):
    pd.DataFrame(rows).to_excel(dir_path / filename, index=False, engine="openpyxl")


# ──────────────────────────────────────────────────────────────────────────────
# Fixtures sintéticas — TRICEL DISTRITO-01.xlsx
# ──────────────────────────────────────────────────────────────────────────────

def _write_tricel_excel(path, *, electos, candidatos, mesa_a_mesa=None):
    """
    electos    : list[(nombre, votos_final)]
    candidatos : list[(num_tricel, nombre, partido, votos)]
    mesa_a_mesa: dict opcional {"columns": [...], "rows": [[...], ...]}
                 — si se omite, se genera una hoja mínima vacía (basta
                 para import_proclamations, que no la lee).

    Reproduce la forma exacta encontrada en DISTRITO-01.xlsx real:
    ELECTOS con encabezado en la fila física 7 (índice 6), CANDIDATOS
    con encabezado en la fila física 6 (índice 5).
    """
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "ELECTOS"
    for _ in range(6):
        ws.append([None, None, None, None])
    ws.append([None, None, "CANDIDATOS ELECTOS", "TOTALES"])
    for nombre, votos in electos:
        ws.append([None, None, nombre, votos])

    ws2 = wb.create_sheet("CANDIDATOS")
    for _ in range(5):
        ws2.append([None] * 6)
    ws2.append([None, "CANDIDATOS", None, "PARTIDO POLÍTICO", "TOTALES", None])
    for num, nombre, partido, votos in candidatos:
        ws2.append([None, num, nombre, partido, votos, None])

    ws3 = wb.create_sheet("MESA A MESA")
    if mesa_a_mesa is not None:
        ws3.append(mesa_a_mesa["columns"])
        for row in mesa_a_mesa["rows"]:
            ws3.append(row)
    else:
        ws3.append(["Región"])

    wb.save(path)


def _default_tricel_fixture(tmp_path, filename="DISTRITO-01.xlsx"):
    _write_tricel_excel(
        tmp_path / filename,
        electos=[("Candidato Uno", 1000)],
        candidatos=[
            (1, "Candidato Uno", "Partido Socialista", 1000),
            (2, "Candidato Dos", "Renovación Nacional", 500),
        ],
    )


# ──────────────────────────────────────────────────────────────────────────────
# import_proclamations
# ──────────────────────────────────────────────────────────────────────────────

class TestImportProclamations:

    def test_candidates_unmatched_vacio_cuando_cruce_es_exacto(self, tmp_path):
        servel_dir = tmp_path / "servel"
        servel_dir.mkdir()
        _write_servel_excel(servel_dir, _raw_servel_rows())

        tricel_dir = tmp_path / "tricel"
        tricel_dir.mkdir()
        _default_tricel_fixture(tricel_dir)

        df, provenance = _import_proclamations_with_servel_dir(tricel_dir, servel_dir)

        assert provenance["candidates_unmatched"] == []
        assert provenance["candidates_matched"] == 1

    def test_schema_canonico_columnas_exactas(self, tmp_path):
        servel_dir = tmp_path / "servel"
        servel_dir.mkdir()
        _write_servel_excel(servel_dir, _raw_servel_rows())

        tricel_dir = tmp_path / "tricel"
        tricel_dir.mkdir()
        _default_tricel_fixture(tricel_dir)

        df, _ = _import_proclamations_with_servel_dir(tricel_dir, servel_dir)

        assert list(df.columns) == [f.name for f in dataclasses.fields(ProclamationRecord)]

    def test_candidato_cruzado_tiene_datos_servel_correctos(self, tmp_path):
        servel_dir = tmp_path / "servel"
        servel_dir.mkdir()
        _write_servel_excel(servel_dir, _raw_servel_rows())

        tricel_dir = tmp_path / "tricel"
        tricel_dir.mkdir()
        _default_tricel_fixture(tricel_dir)

        df, _ = _import_proclamations_with_servel_dir(tricel_dir, servel_dir)

        assert len(df) == 1
        fila = df.iloc[0]
        assert fila["candidate_id"] == 1001
        assert fila["district_id"] == 1
        assert fila["votes_final"] == 1000
        assert bool(fila["officially_elected"]) is True
        assert fila["party_id"] == "ps"

    def test_candidato_sin_cruce_va_a_unmatched(self, tmp_path):
        servel_dir = tmp_path / "servel"
        servel_dir.mkdir()
        _write_servel_excel(servel_dir, _raw_servel_rows())

        tricel_dir = tmp_path / "tricel"
        tricel_dir.mkdir()
        _write_tricel_excel(
            tricel_dir / "DISTRITO-01.xlsx",
            electos=[("Nombre Que No Existe En Servel", 999)],
            candidatos=[
                (1, "Candidato Uno", "Partido Socialista", 1000),
                (2, "Candidato Dos", "Renovación Nacional", 500),
            ],
        )

        df, provenance = _import_proclamations_with_servel_dir(tricel_dir, servel_dir)

        assert df.empty
        assert provenance["candidates_unmatched"] == ["Nombre Que No Existe En Servel"]


def _import_proclamations_with_servel_dir(tricel_dir, servel_dir, monkeypatch=None):
    """
    import_proclamations() resuelve el directorio SERVEL vía
    chiledist.domain.data.get_servel_dir() (CHILEDIST_DATA_DIR) — para
    testear sin depender de esa variable de entorno global, se apunta
    temporalmente vía monkeypatch al servel_dir sintético.
    """
    import chiledist.domain.data.tricel as tricel_mod

    original = tricel_mod.get_servel_dir
    tricel_mod.get_servel_dir = lambda: servel_dir
    try:
        return import_proclamations(
            source_dir=tricel_dir, commune_cut_map=_FAKE_COMMUNE_CUT_MAP,
        )
    finally:
        tricel_mod.get_servel_dir = original


# ──────────────────────────────────────────────────────────────────────────────
# ValidationReport.__str__
# ──────────────────────────────────────────────────────────────────────────────

class TestValidationReportStr:

    def test_formato_exacto(self):
        report = ValidationReport(
            election_id="CL-2025-DIP",
            districts_validated=28, districts_total=28,
            seats_validated=155, seats_total=155,
            list_allocations_validated=96, list_allocations_total=96,
            candidate_proclamations=155, candidate_proclamations_total=155,
            ties_requiring_legal_resolution=0,
            source_hashes={"SERVEL": "aaa", "TRICEL": "bbb"},
            status="EXACT_REPRODUCTION",
            discrepancies=[],
            votes_tricel_fallback_count=0,
        )
        assert str(report) == (
            "Election: Diputados 2025\n"
            "Districts validated: 28/28\n"
            "Seats validated: 155/155\n"
            "List allocations: 96/96\n"
            "Candidate proclamations: 155/155\n"
            "Ties requiring legal resolution: 0\n"
            "Votes with SERVEL fallback: 0/1096\n"
            "Source hashes:\n"
            "  SERVEL: aaa\n"
            "  TRICEL: bbb\n"
            "Status: EXACT_REPRODUCTION"
        )


# ──────────────────────────────────────────────────────────────────────────────
# validate_election
# ──────────────────────────────────────────────────────────────────────────────

def _synthetic_candidates_servel():
    return pd.DataFrame([
        {"district_id": 1, "candidate_id": 101, "candidate_name": "A",
         "party_id": "udi", "list_id": "Pacto1", "votes": 1000},
        {"district_id": 1, "candidate_id": 102, "candidate_name": "B",
         "party_id": "udi", "list_id": "Pacto1", "votes": 200},
        {"district_id": 1, "candidate_id": 103, "candidate_name": "C",
         "party_id": "ps", "list_id": "Pacto2", "votes": 900},
        {"district_id": 1, "candidate_id": 104, "candidate_name": "D",
         "party_id": "ps", "list_id": "Pacto2", "votes": 50},
    ])


_ASSIGNMENT = {"CUT1": 1}
_MAGNITUDES = {1: 2}
_PACTO_MAP = {"udi": "Pacto1", "ps": "Pacto2"}


class TestValidateElection:

    def test_reproduccion_exacta(self):
        proclamations = pd.DataFrame([
            {"district_id": 1, "candidate_id": 101, "candidate_name": "A",
             "party_id": "udi", "list_id": "Pacto1", "votes_final": 1000,
             "officially_elected": True},
            {"district_id": 1, "candidate_id": 103, "candidate_name": "C",
             "party_id": "ps", "list_id": "Pacto2", "votes_final": 900,
             "officially_elected": True},
        ])
        report = validate_election(
            _synthetic_candidates_servel(), proclamations,
            _ASSIGNMENT, _MAGNITUDES, _PACTO_MAP,
        )
        assert report.status == "EXACT_REPRODUCTION"
        assert report.discrepancies == []
        assert report.districts_validated == report.districts_total == 1
        assert report.candidate_proclamations == report.candidate_proclamations_total == 2

    def test_una_discrepancia_resulta_en_partial(self):
        """
        TRICEL proclama a B (200 votos) en vez de A (1000 votos) — el
        motor D'Hondt calcula A como ganador. Un solo escaño mal
        asignado produce 2 entradas de discrepancia (el electo TRICEL
        que el motor no calculó + el electo del motor que TRICEL no
        proclamó): son las dos caras del mismo desacuerdo candidato a
        candidato.
        """
        proclamations = pd.DataFrame([
            {"district_id": 1, "candidate_id": 102, "candidate_name": "B",
             "party_id": "udi", "list_id": "Pacto1", "votes_final": 200,
             "officially_elected": True},
            {"district_id": 1, "candidate_id": 103, "candidate_name": "C",
             "party_id": "ps", "list_id": "Pacto2", "votes_final": 900,
             "officially_elected": True},
        ])
        report = validate_election(
            _synthetic_candidates_servel(), proclamations,
            _ASSIGNMENT, _MAGNITUDES, _PACTO_MAP,
        )
        assert report.status == "PARTIAL"
        assert len(report.discrepancies) == 2
        assert all(d["type"] == "candidate_mismatch" for d in report.discrepancies)
        # a nivel de lista (pacto), el conteo de escaños sigue siendo correcto:
        # la discrepancia es de QUIÉN gana el escaño, no CUÁNTOS gana cada pacto
        assert report.list_allocations_validated == report.list_allocations_total == 2
        assert report.candidate_proclamations == 1
        assert report.candidate_proclamations_total == 2


# ──────────────────────────────────────────────────────────────────────────────
# Infraestructura — _paths / .env.example
# ──────────────────────────────────────────────────────────────────────────────

class TestDataPaths:

    def test_get_servel_dir_honra_chiledist_data_dir(self, monkeypatch):
        monkeypatch.setenv("CHILEDIST_DATA_DIR", "/tmp/fake_data_root")
        assert str(get_servel_dir()) == os.path.join("/tmp/fake_data_root", "SERVEL_2025")

    def test_get_tricel_dir_honra_chiledist_data_dir(self, monkeypatch):
        monkeypatch.setenv("CHILEDIST_DATA_DIR", "/tmp/fake_data_root")
        assert str(get_tricel_dir()) == os.path.join("/tmp/fake_data_root", "TRICEL_2025")

    def test_get_tricel_dir_usa_default_sin_env_var(self, monkeypatch):
        monkeypatch.delenv("CHILEDIST_DATA_DIR", raising=False)
        assert str(get_tricel_dir()) == os.path.join("datos/fuentes", "TRICEL_2025")


class TestEnvExample:

    def test_env_example_existe(self):
        root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        assert os.path.exists(os.path.join(root, ".env.example"))

    def test_env_example_documenta_chiledist_data_dir(self):
        root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        with open(os.path.join(root, ".env.example"), encoding="utf-8") as f:
            content = f.read()
        assert "CHILEDIST_DATA_DIR" in content
